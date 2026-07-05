from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from datetime import datetime, timedelta, date
from bot.database.connection import get_db
from bot.database import crud
from bot.keyboards.reply import get_admin_keyboard, get_main_keyboard
from bot.config import ADMIN_TELEGRAM_IDS, get_current_date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io
import asyncio
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

router = Router()


class BroadcastState(StatesGroup):
    waiting_for_message = State()


EXCEL_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


async def is_admin(telegram_id: int) -> bool:
    if telegram_id in ADMIN_TELEGRAM_IDS:
        return True
    async with get_db() as session:
        admin = await crud.get_admin_by_telegram_id(session, telegram_id)
        return admin is not None


@router.message(Command("admin"))
async def admin_panel(message: Message):
    if not await is_admin(message.from_user.id):
        await message.answer("⛔ Siz admin emassiz.\n\nBu buyruq faqat adminlar uchun.")
        return

    await message.answer(
        "🔑 <b>ADMIN PANELI</b>\n\n"
        "Quyidagi hisobot turlaridan birini tanlang:\n\n"
        "📊 <b>Kunlik hisobot</b> — bugungi buyurtmalar\n"
        "📊 <b>Haftalik hisobot</b> — shu haftadagi buyurtmalar\n"
        "📊 <b>Oylik hisobot</b> — shu oydagi buyurtmalar\n"
        "👥 <b>Barcha mijozlar</b> — Excel formatda\n"
        "📢 <b>Xabar yuborish</b> — Barcha foydalanuvchilarga",
        reply_markup=get_admin_keyboard()
    )


@router.message(F.text == "◀️ Chiqish")
async def admin_exit(message: Message):
    await message.answer("🏠 <b>Bosh menyu</b>", reply_markup=get_main_keyboard())


async def get_report_data(session, start_date: date, end_date: date):
    orders = await crud.get_orders_by_date_range(session, start_date, end_date)
    users = await crud.get_all_users(session)
    user_map = {u.id: u for u in users}

    rows = []
    for order in orders:
        user = user_map.get(order.user_id)
        if user:
            rows.append({
                "order_number": order.order_number,
                "full_name": user.full_name,
                "phone": user.phone,
                "household_size": user.household_size,
                "quantity": order.quantity,
                "address": order.address,
                "created_at": order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "",
                "status": "Yetkazilgan" if order.status.name == "delivered" else "Kutilmoqda"
            })
    return rows, len(orders)


def _style_excel_header(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = EXCEL_HEADER_FONT
        cell.fill = EXCEL_HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_excel_row(ws, row, values, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = values[col - 1]
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')


def create_excel(rows, start_label: str, end_label: str, title: str = "Hisobot", extra_cols: list = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.merge_cells(f'A1:G1')
    cell = ws['A1']
    cell.value = f"OKS Suv — {title} ({start_label} — {end_label})"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ["№", "Mijoz ismi", "Telefon", "Xonadon (kishi)", "Soni (19L)", "Buyurtma vaqti", "Holat"]
    _style_excel_header(ws, 3, headers)

    for i, row in enumerate(rows, 1):
        values = [i, row["full_name"], row["phone"], row["household_size"],
                  row["quantity"], row["created_at"], row["status"]]
        _style_excel_row(ws, i + 3, values, len(headers))

    widths = [5, 25, 18, 18, 14, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_pdf(rows, start_label: str, end_label: str, total: int):
    if not HAS_REPORTLAB:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("OKS Suv — Hisobot", styles['Title']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>{start_label}</b> — <b>{end_label}</b>", styles['Normal']))
    elements.append(Paragraph(f"Jami buyurtmalar: <b>{total}</b>", styles['Normal']))
    elements.append(Spacer(1, 10))

    table_data = [["№", "Mijoz ismi", "Telefon", "Xonadon", "Soni", "Vaqt", "Holat"]]
    for i, row in enumerate(rows, 1):
        table_data.append([str(i), row["full_name"], row["phone"],
                          str(row["household_size"]), str(row["quantity"]),
                          row["created_at"], row["status"]])

    page_width = A4[0] - 30*mm
    col_widths = [page_width * 0.06, page_width * 0.22, page_width * 0.16,
                  page_width * 0.12, page_width * 0.10, page_width * 0.18, page_width * 0.16]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf


async def send_report(message: Message, rows, total, start_label: str, end_label: str, report_type: str):
    if not rows:
        await message.answer(
            f"📊 <b>{report_type}</b>\n\n"
            f"<b>{start_label}</b> dan <b>{end_label}</b> gacha\n\n"
            "📭 Buyurtmalar mavjud emas."
        )
        return

    await message.answer(
        f"📊 <b>{report_type}</b>\n\n"
        f"📅 <b>{start_label}</b> — <b>{end_label}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📦 Jami buyurtmalar: <b>{total}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n\n"
        f"📎 Quyida Excel va PDF formatlarini yuklab oling:"
    )

    excel_buf = create_excel(rows, start_label, end_label, report_type)
    sanitized = report_type.lower().replace(' ', '_')
    await message.answer_document(
        BufferedInputFile(file=excel_buf.read(), filename=f"{sanitized}_{end_label}.xlsx"),
        caption="📊 <b>Excel format</b>"
    )

    pdf_buf = create_pdf(rows, start_label, end_label, total)
    if pdf_buf:
        await message.answer_document(
            BufferedInputFile(file=pdf_buf.read(), filename=f"{sanitized}_{end_label}.pdf"),
            caption="📄 <b>PDF format</b>"
        )


@router.message(F.text == "📊 Kunlik hisobot")
async def daily_report(message: Message):
    if not await is_admin(message.from_user.id):
        return
    today = get_current_date()
    async with get_db() as session:
        rows, total = await get_report_data(session, today, today)
    await send_report(message, rows, total,
                      today.strftime('%d.%m.%Y'), today.strftime('%d.%m.%Y'),
                      "Kunlik hisobot")


@router.message(F.text == "📊 Haftalik hisobot")
async def weekly_report(message: Message):
    if not await is_admin(message.from_user.id):
        return
    today = get_current_date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    async with get_db() as session:
        rows, total = await get_report_data(session, start_of_week, end_of_week)
    await send_report(message, rows, total,
                      start_of_week.strftime('%d.%m.%Y'), end_of_week.strftime('%d.%m.%Y'),
                      "Haftalik hisobot")


@router.message(F.text == "📊 Oylik hisobot")
async def monthly_report(message: Message):
    if not await is_admin(message.from_user.id):
        return
    today = get_current_date()
    start_of_month = today.replace(day=1)
    if today.month == 12:
        end_of_month = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_of_month = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    async with get_db() as session:
        rows, total = await get_report_data(session, start_of_month, end_of_month)
    month_names = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
        7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
    }
    await send_report(message, rows, total,
                      f"{start_of_month.strftime('%d.%m.%Y')}", f"{end_of_month.strftime('%d.%m.%Y')}",
                      f"Oylik hisobot ({month_names.get(today.month, '')})")


@router.message(F.text == "👥 Barcha mijozlar (Excel)")
async def all_clients_excel(message: Message):
    if not await is_admin(message.from_user.id):
        return

    async with get_db() as session:
        users = await crud.get_all_users(session)

    if not users:
        await message.answer("👥 Mijozlar ro'yxati bo'sh.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mijozlar"

    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "OKS Suv — Barcha mijozlar"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ["№", "Ism", "Telefon", "Xonadon (kishi)", "Ro'yxatdan o'tgan"]
    _style_excel_header(ws, 3, headers)

    for i, user in enumerate(users, 1):
        values = [i, user.full_name, user.phone, user.household_size,
                  user.created_at.strftime("%d.%m.%Y") if user.created_at else ""]
        _style_excel_row(ws, i + 3, values, 5)

    widths = [5, 25, 18, 18, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = get_current_date().strftime('%d.%m.%Y')
    await message.answer_document(
        BufferedInputFile(file=buf.read(), filename=f"barcha_mijozlar_{today}.xlsx"),
        caption=f"👥 <b>Barcha mijozlar</b>\nJami: {len(users)} ta"
    )


# ─── BROADCAST (BARCHA FOYDALANUVCHILARGA XABAR YUBORISH) ───

@router.message(F.text == "📢 Xabar yuborish")
async def broadcast_start(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        return
    
    await message.answer(
        "📢 <b>BARCHA FOYDALANUVCHILARGA XABAR YUBORISH</b>\n\n"
        "Yubormoqchi bo'lgan xabaringizni yozing.\n\n"
        "⚠️ <b>Diqqat:</b>\n"
        "• Xabar <b>HTML formatda</b> bo'lishi mumkin\n"
        "• <b>Bold</b> uchun: &lt;b&gt;matn&lt;/b&gt;\n"
        "• <b>Italic</b> uchun: &lt;i&gt;matn&lt;/i&gt;\n"
        "• Rasm, video, sticker yuborish mumkin\n\n"
        "📝 Xabaringizni yuboring yoki /cancel bosing:",
        reply_markup=get_main_keyboard()
    )
    await state.set_state(BroadcastState.waiting_for_message)


@router.message(BroadcastState.waiting_for_message, Command("cancel"))
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "🚫 Xabar yuborish bekor qilindi.",
        reply_markup=get_admin_keyboard()
    )


@router.message(BroadcastState.waiting_for_message)
async def broadcast_send(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        await state.clear()
        return
    
    # Foydalanuvchilarni olish
    async with get_db() as session:
        users = await crud.get_all_users(session)
    
    if not users:
        await message.answer(
            "❌ Hech qanday foydalanuvchi topilmadi.",
            reply_markup=get_admin_keyboard()
        )
        await state.clear()
        return
    
    # Tasdiqlash xabari
    confirm_msg = await message.answer(
        f"📊 <b>XABAR YUBORISH</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{len(users)}</b>\n\n"
        f"⏳ Xabar yuborilmoqda...\n\n"
        f"Iltimos, kuting..."
    )
    
    # Statistika
    success_count = 0
    failed_count = 0
    blocked_count = 0
    
    # Har bir foydalanuvchiga xabar yuborish
    for user in users:
        try:
            # Xabar turini aniqlash va yuborish
            if message.text:
                await message.bot.send_message(
                    chat_id=user.telegram_id,
                    text=message.text
                )
            elif message.photo:
                await message.bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=message.photo[-1].file_id,
                    caption=message.caption
                )
            elif message.video:
                await message.bot.send_video(
                    chat_id=user.telegram_id,
                    video=message.video.file_id,
                    caption=message.caption
                )
            elif message.document:
                await message.bot.send_document(
                    chat_id=user.telegram_id,
                    document=message.document.file_id,
                    caption=message.caption
                )
            elif message.sticker:
                await message.bot.send_sticker(
                    chat_id=user.telegram_id,
                    sticker=message.sticker.file_id
                )
            elif message.voice:
                await message.bot.send_voice(
                    chat_id=user.telegram_id,
                    voice=message.voice.file_id,
                    caption=message.caption
                )
            elif message.audio:
                await message.bot.send_audio(
                    chat_id=user.telegram_id,
                    audio=message.audio.file_id,
                    caption=message.caption
                )
            elif message.animation:
                await message.bot.send_animation(
                    chat_id=user.telegram_id,
                    animation=message.animation.file_id,
                    caption=message.caption
                )
            else:
                # Noma'lum format
                failed_count += 1
                continue
            
            success_count += 1
            
            # Har 50 ta xabardan keyin kichik pauza (flood oldini olish)
            if success_count % 50 == 0:
                await asyncio.sleep(1)
            else:
                await asyncio.sleep(0.05)  # 50ms pauza
            
        except Exception as e:
            error_str = str(e).lower()
            if "blocked" in error_str or "deactivated" in error_str or "user is deactivated" in error_str:
                blocked_count += 1
            else:
                failed_count += 1
            logger.warning(f"Failed to send message to user {user.telegram_id}: {e}")
            continue
    
    # Natijani ko'rsatish
    await confirm_msg.edit_text(
        f"✅ <b>XABAR YUBORILDI!</b>\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📊 <b>STATISTIKA:</b>\n\n"
        f"👥 Jami: <b>{len(users)}</b>\n"
        f"✅ Yuborildi: <b>{success_count}</b>\n"
        f"❌ Xato: <b>{failed_count}</b>\n"
        f"🚫 Bloklangan: <b>{blocked_count}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n\n"
        f"💡 Yuborilish foizi: <b>{(success_count / len(users) * 100):.1f}%</b>"
    )
    
    await message.answer(
        "🏠 Admin paneliga qaytish uchun /admin bosing.",
        reply_markup=get_admin_keyboard()
    )
    
    await state.clear()
